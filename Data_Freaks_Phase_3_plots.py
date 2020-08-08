import pandas as pd
import matplotlib.pyplot as plt
from wordcloud import WordCloud
import re


dfs = pd.read_excel('sentiment212.xlsx')
new_df = dfs.dropna()


def create_country_cloud():
    loc_count = new_df.groupby(['location']).size()
    loc_list = loc_count.to_dict()

    wordcloud = WordCloud(background_color="#bce8bf", margin=0).generate_from_frequencies(frequencies=loc_list)

    # Display the generated image:
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")
    plt.title("Word Cloud based on Frequency")
    plt.margins(x=0, y=0)
    plt.show()


def write_text():
    text = new_df['text'].tolist()
    file = open("testfile.txt", "w", encoding='utf-8')

    for i in range(len(text)):
        file.write(text[i])
    file.close()


def word_cloud():
    frequency = {}
    openFile = open("testfile.txt", 'r', encoding="utf8")
    fileString = openFile.read()
    all_words = re.findall(r'(\b[A-Za-z][a-z]{2,9}\b)', fileString)

    words = []
    for word in all_words:
        words.append(word.lower())

    find_words = ['anxiety', 'stress', 'bipolar', 'disorder', 'troubling', 'attack', 'uncertain', 'depression',
                  'pain', 'mental', 'health', 'issues', 'battles', 'struggling', 'disturbing', 'suffering',
                  'depressed', 'nervous', 'panic', 'suffered', 'anxious', 'unhappy', 'terrible', 'painful', 'suicidal',
                  'broken', 'worsening', 'harmed', 'tired', 'died', 'damage', 'trapped', 'stress', 'lonliness',
                  'rescue', 'die', 'toxic', 'bullying', 'failure', 'suicide', 'risks', 'kill', 'hard', 'prisoner',
                  'rescue', 'worse', 'worst', 'please', 'horrible', 'crippling', 'vomit', 'psychology', 'bitter',
                  'beat', 'pills', 'manic', 'not', 'extreme', 'experience', 'weakness', 'confession', 'blames',
                  'enemies', 'obsessive', 'threats', 'terror', 'judgment', 'crazy', 'bullies', 'abuse', 'survivors',
                  'dead', 'trouble', 'victims', 'adverse']

    for word in words:
        count = frequency.get(word, 0)
        frequency[word] = count + 1

    new_freq = {k: frequency[k] for k in find_words if k in frequency}

    wordcloud = WordCloud(background_color="black", margin=0).generate_from_frequencies(frequencies=new_freq)

    # Display the generated image:
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")
    plt.title("Word Cloud based on Suicidal words")
    plt.margins(x=0, y=0)
    plt.show()


def show_country_data():
    country = []
    country_freq = []
    loc_count = new_df.groupby(['location']).size()
    loc_list = loc_count.to_dict()
    for key, value in loc_list.items():
        if 'USA' in key or 'usa' in key or 'Canada' in key \
                or 'India' in key or 'Australia' in key or 'England' in key:
            country.append(key)
            country_freq.append(value)

    colours = range(len(country))
    plt.scatter(country, country_freq, c=colours)
    plt.xticks(rotation ='vertical', size='xx-small', linespacing=30.0)
    plt.xlabel('Country')
    plt.ylabel('Tweets Tweeted per country')

    plt.show()


def user_id_follower_count():
    import openpyxl
    import matplotlib.pyplot as plt

    path = "sentiment212.xlsx"

    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    userid_list = []
    followerscount_list = []
    for i in range(1, m_row + 1):
        user_id = sheet_obj.cell(row=i, column=10)
        # print(user_id.value)
        if user_id.value != 'NA':
            userid_list.append(str(user_id.value))

    for i in range(1, m_row + 1):
        followers_count = sheet_obj.cell(row=i, column=11)
        # print(followers_count.value)
        if followers_count.value != 'NA':
            followerscount_list.append(followers_count.value)

    userid_list.remove('1058753888')
    userid_list.remove('350597180')
    userid_list.remove('2957306475')
    userid_list.remove('849904956047974401')
    userid_list.remove('754093423')
    userid_list.remove('701480597865283584')
    userid_list.remove('85725864')
    userid_list.remove('716202844219699200')
    userid_list.remove('47535811')
    followerscount_list.remove(6977554)
    followerscount_list.remove(977281)
    followerscount_list.remove(568917)
    followerscount_list.remove(565303)
    followerscount_list.remove(432155)
    followerscount_list.remove(354009)
    followerscount_list.remove(318434)
    followerscount_list.remove(288305)
    followerscount_list.remove(201571)

    """print("follower count list: ", max(followerscount_list[1:]))
    print("index: ", followerscount_list[1:].index(201571)) #1430
    print("userid: ", userid_list[1430]) #1058753888"""

    plt.bar(userid_list[1:15], followerscount_list[1:15])
    plt.xlabel('User id', fontsize=5)
    plt.ylabel('Follower count', fontsize=5)
    plt.xticks(rotation='vertical', size='xx-small')
    plt.show()

def pairwise_comp():
    import pandas as pd
    from matplotlib import pyplot as plt
    import seaborn as sns

    df = pd.read_excel('sentiment212.xlsx', sheet_name='Sheet1')

    sns.pairplot(data=df[['friends_count', 'statuses_count', 'userid', 'followers_count']])
    plt.show()


def main():
    write_text()
    word_cloud()
    show_country_data()
    user_id_follower_count()
    pairwise_comp()


if __name__ == '__main__':
    main()